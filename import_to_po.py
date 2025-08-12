import polib
import config
from openpyxl import load_workbook
from datetime import datetime

# Load Excel file
wb = load_workbook(config.IMPORTED_XLSX_FILE)
ws = wb.active

# Create PO file with metadata
po = polib.POFile()

po.metadata = {
    'Project-Id-Version': '1.0',
    'Report-Msgid-Bugs-To': '',
    'POT-Creation-Date': datetime.now().strftime("%Y-%m-%d %H:%M%z"),
    'PO-Revision-Date': datetime.now().strftime("%Y-%m-%d %H:%M%z"),
    'Last-Translator': 'Miguel Sanchez <miguelangel.sanchez@novaigrup.com>',
    'Language-Team': 'Portuguese',
    'Language': 'pt',
    'MIME-Version': '1.0',
    'Content-Type': 'text/plain; charset=UTF-8',
    'Content-Transfer-Encoding': '8bit',
}

# Read translations from XLSX
for row in ws.iter_rows(min_row=2, values_only=True):
    msgid, msgstr = row
    if msgid is None:
        continue  # Skip empty rows
    
    entry = polib.POEntry(
        msgid=str(msgid),
        msgstr=str(msgstr) if msgstr else ""
    )
    po.append(entry)

# Save PO file
po.save(config.GENERATED_PO_FILE)

print(f"PO file generated: {config.GENERATED_PO_FILE}")
